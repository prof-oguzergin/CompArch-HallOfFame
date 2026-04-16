from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data = {
    "HPCA": [
        (1995,190,36),(1996,131,29),(1997,152,30),(1998,141,26),(1999,117,37),
        (2000,163,35),(2001,110,26),(2002,130,26),(2003,141,31),(2004,153,27),
        (2005,181,28),(2006,172,22),(2007,174,28),(2008,161,31),(2009,184,35),
        (2010,175,32),(2011,227,42),(2012,213,36),(2013,249,51),(2014,215,55),
        (2015,228,51),(2016,240,53),(2017,224,50),(2018,260,54),(2019,233,46),
        (2020,248,48),(2021,260,61),(2022,273,80),(2023,364,91),(2024,410,75),
        (2025,534,112),(2026,602,119),
    ],
    "MICRO": [
        (2009,209,52),(2010,248,42),(2011,209,44),(2012,228,40),(2013,239,39),
        (2014,279,53),(2015,283,61),(2016,288,61),(2017,327,61),(2018,351,74),
        (2019,345,79),(2020,446,82),(2021,430,94),(2022,366,83),(2023,434,101),
        (2024,497,113),
    ],
    "ISCA": [
        (2001,163,24),(2002,180,27),(2003,184,36),(2004,217,31),(2005,194,45),
        (2006,234,30),(2007,204,46),(2008,259,37),(2009,210,43),(2010,245,44),
        (2011,208,40),(2012,262,47),(2013,288,56),(2014,258,46),(2015,305,58),
        (2016,291,57),(2017,322,54),(2018,367,64),(2019,365,62),(2020,428,77),
        (2021,407,76),(2022,400,67),(2023,372,79),(2024,423,83),(2025,570,132),
    ],
    "ASPLOS": [
        (2002,146,29),(2003,109,25),(2004,123,28),(2005,114,24),(2006,175,24),
        (2007,169,24),(2008,158,38),(2009,127,31),(2010,113,29),(2011,181,32),
        (2012,152,32),(2013,193,44),(2014,217,49),(2015,287,48),(2016,232,53),
        (2017,320,53),(2018,319,56),(2019,351,74),(2020,486,86),(2021,398,75),
        (2022,397,80),(2023,600,128),(2024,922,193),(2025,912,160),
    ],
}

wb = Workbook()
wb.remove(wb.active)

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="center", vertical="center")

data_font = Font(name="Arial", size=10)
data_align_center = Alignment(horizontal="center")
data_align_right = Alignment(horizontal="right")

alt_fill = PatternFill("solid", start_color="D6E4F0")

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [8, 12, 12, 12]

for venue, rows in data.items():
    ws = wb.create_sheet(venue)
    headers = ["Year", "Submitted", "Accepted", "Rate (%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 20

    for r_idx, (year, sub, acc) in enumerate(rows, 2):
        rate = round(acc / sub * 100, 1)
        values = [year, sub, acc, rate]
        fill = alt_fill if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = border
            if fill:
                cell.fill = fill
            if c_idx == 1:
                cell.alignment = data_align_center
            elif c_idx == 4:
                cell.number_format = '0.0"%"'
                cell.alignment = data_align_right
            else:
                cell.alignment = data_align_right

    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

out_path = r"C:\Users\Z GAMES\Yapay Zeka\CompArch-HallOfFame\acceptance_rates.xlsx"
wb.save(out_path)
print("Saved:", out_path)
