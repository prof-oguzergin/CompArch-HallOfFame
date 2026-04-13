"""Generate Top Picks Excel from data.js (all 22 years: 2003-2024)"""
import re, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# Parse data.js to extract toppicks_papers
with open("data.js", encoding="utf-8") as f:
    content = f.read()

# Extract toppicks_papers array using regex (ends with ]\n, before toppicks:[])
match = re.search(r'toppicks_papers:\s*\[(.*?)\n\]', content, re.DOTALL)
if not match:
    raise ValueError("Could not find toppicks_papers in data.js")

raw = match.group(1)

# Parse each paper entry
papers = []
for m in re.finditer(
    r'\{year:(\d+),type:"(TP|HM)",conf:"([^"]*)",title:"([^"]*)",authors:\[([^\]]*)\]\}',
    raw
):
    year = int(m.group(1))
    ptype = "Top Pick" if m.group(2) == "TP" else "Honorable Mention"
    conf = m.group(3)
    title = m.group(4)
    authors_raw = m.group(5)
    authors = [a.strip().strip('"') for a in authors_raw.split('","')] if authors_raw.strip() else []
    papers.append({
        "year": year, "type": ptype, "conf": conf,
        "title": title, "authors": authors
    })

print(f"Parsed {len(papers)} papers ({sum(1 for p in papers if p['type']=='Top Pick')} TP, {sum(1 for p in papers if p['type']=='Honorable Mention')} HM)")

# IEEE Micro issue mapping
def ieee_issue(year):
    vol = 24 + (year - 2003)
    return f"vol.{vol} {year+1}"

wb = Workbook()
header_font = Font(bold=True, size=11, name="Arial")
header_fill = PatternFill("solid", fgColor="F59E0B")
tp_fill = PatternFill("solid", fgColor="E8F5E9")
hm_fill = PatternFill("solid", fgColor="FFF3E0")
thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

# Sheet 1: All Papers
ws = wb.active
ws.title = "Top Picks Papers"

headers = ["Year", "Title", "Authors", "Conference", "Type", "IEEE Micro Issue"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

# Sort by year descending, then type (TP first), then title
sorted_papers = sorted(papers, key=lambda p: (-p["year"], 0 if p["type"]=="Top Pick" else 1, p["title"]))

for i, p in enumerate(sorted_papers, 2):
    authors_str = ", ".join(p["authors"]) if p["authors"] else ""
    row_data = [p["year"], p["title"], authors_str, p["conf"], p["type"], ieee_issue(p["year"])]
    for j, val in enumerate(row_data, 1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = thin_border
        c.fill = tp_fill if p["type"] == "Top Pick" else hm_fill

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 80
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.freeze_panes = 'A2'

# Sheet 2: Author Summary (TP only)
ws2 = wb.create_sheet("Author Summary")
author_tp = defaultdict(int)
author_years = defaultdict(set)

for p in papers:
    if p["type"] != "Top Pick":
        continue
    for a in p["authors"]:
        author_tp[a] += 1
        author_years[a].add(p["year"])

summary = []
for a in author_tp:
    tp = author_tp[a]
    years = sorted(author_years[a])
    span = f"{min(years)}-{max(years)}" if len(years) > 1 else str(years[0])
    summary.append((a, tp, ", ".join(str(y) for y in years), span))
summary.sort(key=lambda x: -x[1])

headers2 = ["Author", "Top Picks", "Years", "Span"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

for i, (name, tp, years, span) in enumerate(summary, 2):
    ws2.cell(row=i, column=1, value=name).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=2, value=tp).font = Font(name="Arial", size=10, bold=True)
    ws2.cell(row=i, column=3, value=years).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=4, value=span).font = Font(name="Arial", size=10)
    for j in range(1, 5):
        ws2.cell(row=i, column=j).border = thin_border

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 12
ws2.freeze_panes = 'A2'

# Sheet 3: Conference-Year Breakdown
ws3 = wb.create_sheet("Conference-Year Breakdown")
conf_year = defaultdict(lambda: defaultdict(int))
all_confs = set()
all_years = sorted(set(p["year"] for p in papers if p["type"]=="Top Pick"))

for p in papers:
    if p["type"] != "Top Pick":
        continue
    conf_base = re.sub(r'\s*\d{4}$', '', p["conf"])
    all_confs.add(conf_base)
    conf_year[conf_base][p["year"]] += 1

confs_sorted = sorted(all_confs, key=lambda c: -sum(conf_year[c].values()))

ws3.cell(row=1, column=1, value="Conference").font = header_font
ws3.cell(row=1, column=1).fill = header_fill
for j, y in enumerate(all_years, 2):
    c = ws3.cell(row=1, column=j, value=y)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
ws3.cell(row=1, column=len(all_years)+2, value="Total").font = header_font
ws3.cell(row=1, column=len(all_years)+2).fill = header_fill

for i, conf in enumerate(confs_sorted, 2):
    ws3.cell(row=i, column=1, value=conf).font = Font(name="Arial", size=10, bold=True)
    total = 0
    for j, y in enumerate(all_years, 2):
        v = conf_year[conf].get(y, 0)
        total += v
        c = ws3.cell(row=i, column=j, value=v if v else "")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
    ws3.cell(row=i, column=len(all_years)+2, value=total).font = Font(name="Arial", size=10, bold=True)

# Total row
r = len(confs_sorted) + 2
ws3.cell(row=r, column=1, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
grand = 0
for j, y in enumerate(all_years, 2):
    v = sum(conf_year[c].get(y, 0) for c in confs_sorted)
    grand += v
    ws3.cell(row=r, column=j, value=v).font = Font(name="Arial", size=10, bold=True)
ws3.cell(row=r, column=len(all_years)+2, value=grand).font = Font(name="Arial", size=10, bold=True)

ws3.column_dimensions['A'].width = 18
for j in range(2, len(all_years)+3):
    ws3.column_dimensions[chr(64+j) if j <= 26 else 'A' + chr(64+j-26)].width = 7
ws3.freeze_panes = 'B2'

out = r"G:\My Drive\Claude Code\Top picks\TopPicks_HallOfFame.xlsx"
wb.save(out)
print(f"Saved to {out} ({len(summary)} authors, {len(sorted_papers)} papers, {len(all_years)} years)")
